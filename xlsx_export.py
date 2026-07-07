from openpyxl import Workbook
from openpyxl.styles import PatternFill
from colorama import Fore, init

init(autoreset=True)

XANH_LA_NHAT = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # tên
XANH_LA = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")       # >=8
XANH_DUONG = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")    # >=6.5
VANG = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")          # >=5
DO = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")


def mau_theo_diem(diem):
    if diem >9:
        return XANH_LA
    elif diem <9:
        return XANH_DUONG
    elif diem <7:
        return VANG
    elif diem <5:
        return DO
    
def xuat_file(danhsach, ten_file="bang_diem.xlxs"):
    tat_ca_mon = set()
    for hoc_sinh in danhsach:
        for mon in hoc_sinh.mon_hoc:
            tat_ca_mon.add(mon["ten"])
        tat_ca_mon = list(tat_ca_mon)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Bảng Điểm"
        
        header = ["ID", "Tên", "Lớp", "SĐT_PH", "Xếp_Loại"] + tat_ca_mon
        
        for vi_tri, ten_cot in enumerate(header):
            ws.cell(row = 1, column= vi_tri + 1, value=ten_cot)
        
        for gia_tri_row, hoc_sinh in enumerate(danhsach):
            row = gia_tri_row + 2
        
        diem_cua_hoc_sinh = {}
        for mon in hoc_sinh.danhsach:
            diem_cua_hoc_sinh[mon["ten"]] = mon["diem"]
        
        fields = [hoc_sinh.id, hoc_sinh.name, hoc_sinh.lop, hoc_sinh.sdt_cha_me, hoc_sinh.xep_loai()]
        
        
        for vitri, thong_tin in enumerate(fields):
            cell = ws.cell(row=row, column=vitri+6, value= thong_tin)
            if vitri == 1:
                cell.fill = XANH_LA_NHAT
                
                
        for vitri, ten_mon in enumerate(tat_ca_mon):
            column = vitri + 6
            diem = diem_cua_hoc_sinh.get(ten_mon)
            cell = ws.cell(row = row, column=vitri+6, value=ten_mon)
            if diem is not None:
                cell.fill = mau_theo_diem(diem)